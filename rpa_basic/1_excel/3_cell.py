from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "NadoSheet"

ws['A1'] = 1
ws['A2'] = 2
ws['A3'] = 3

ws['B1'] = 6
ws['B2'] = 4
ws['B3'] = 5

print(ws["A1"])         # A1 셀의 정보를 출력
print(ws["A1"].value)   # A1 셀의 '값'을 출력
print(ws["A10"].value)  # 값이 없을 땐 'None'을 출력

print(ws.cell(row=1, column=1).value)    # A1 print(ws["A1"].value) 이걸아 같음 

c = ws.cell(column=3, row=1, value=10)  # ws["C1"].value = 10, ws["C1"] = 10 과 동일하다
print(c.value)  # ws["C1"].value

from random import *

for x in range(1, 11):
    for y in range(1, 11):
        ws.cell(row=x, column=y, value=randint(0, 100))

wb.save("sample.xlsx")
