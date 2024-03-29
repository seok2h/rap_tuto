from openpyxl import load_workbook


# wb = load_workbook("sample_formula.xlsx")
# ws = wb.active

# # print(ws.values)

# # 수식 그대로 가져오고 있음
# for row in ws.values:
#     for cell in row:
#         print(cell)


# 수식이 아닌 실제 데이터를 가지고 옴
# evaluate 되지 않은 상태의 데이터는 None이라고 표시

wb = load_workbook("sample_formula.xlsx", data_only=True)
ws = wb.active

for row in ws.values:
    for cell in row:
        print(cell)