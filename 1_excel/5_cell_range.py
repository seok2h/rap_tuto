from openpyxl import Workbook
from random import *
wb = Workbook()
ws = wb.active

# 1줄씩 데이터 넣기
ws.append((["번호", "영어", "수학"]))
for i in range(1, 11):
    ws.append([i, randint(0, 100), randint(0, 100)])

# col_B = ws["B"] # 영어 column만 가져오기
# print(col_B)
# for cell in col_B:
#     print(cell.value)


# col_range = ws["B:C"] # 영어, 수학 column 함께 가지고 오기
# for cols in col_range:
#     for cell in cols:
#         print(cell.value)

row_title = ws[1] # 1번째 row만 가지고 오기
# for cell in row_title:
#     print(cell.value)

# row_range = ws[2:6] # 2번째 줄에서 6번째 줄까지 가지고 오기
# for rows in row_range:
#     for cell in rows:
#         print(cell.value, end=" ")
#     print()

# from openpyxl.utils.cell import coordinate_from_string

# row_range = ws[2:ws.max_row]  # 몇 번째 줄까지 있는지 모를 때 사용 가능
# for rows in row_range:
#     for cell in rows:
#         # print(cell.value, end=" ")
#         # print(cell.coordinate, end=" ")     # A10, AZ250
#         xy = coordinate_from_string(cell.coordinate)  # ('A', 2) 이것처럼 튜플 형태로 저장된다.
#         # print(xy, end=" ")
#         print(xy[0], end="")
#         print(xy[1], end=" ")
#     print()

# 전체 rows
# print(tuple(ws.rows)) 

# for row in tuple(ws.rows):
#     print(row[1].value)

# 전체 columns
# print(tuple(ws.columns))
# for column in tuple(ws.columns):
#     print(column[0].value)

# for row in ws.iter_rows():  # 전체 row
#     print(row[2].value)

# for row in ws.iter_cols():  # 전체 column
#     print(column[0].value)

# 2번째 줄부터 11번째 줄까지, 2번째 열부터 3번째 열까지
for row in ws.iter_rows(min_row=2, max_row=11, min_col=2, max_col=3):
    print(row[0].value, row[1].value)  # 수학, 영어
    print(row)

for col in ws.iter_cols(min_row=1, max_row=5, min_col=1, max_col=3):
    print(col)

wb.save("sample.xlsx")

