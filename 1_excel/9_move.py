from openpyxl import load_workbook
wb = load_workbook("sample.xlsx")
ws = wb.active

# 번호 영어 수학
# 번호 (국어) 영어 수학

# ws.move_range("B1:C11", rows=0, cols=1)     #ws.move_range("움직일 범위", 행기준으로 몇 줄, 열 기준으로 몇 줄)
# ws["B1"].value = "국어" # B1 셀에 국어 입력

#번호 영어 수학
ws.move_range("C1:C11", rows=5, cols=-1)    # 수학 영역을 row 기준 밑으로 5칸 col 기준 왼쪽으로 (음수이기 때문에) 1칸 
wb.save("sample_korean.xlsx")