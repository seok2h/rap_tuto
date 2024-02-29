from openpyxl.styles import Border, Font, Side, PatternFill
from openpyxl import load_workbook
wb = load_workbook("sample.xlsx")
ws = wb.active

a1 = ws["A1"]   # 번호
b1 = ws["B1"]   # 영어
c1 = ws["C1"]   # 수학

# A 열의 너비를 5로 지정
ws.column_dimensions["A"].width = 5

# 1 행의 높이를 50으로 설정
ws.row_dimensions[1].width = 50

# 스타일 적용
a1.font = Font(color="FF0000", italic=True, bold=True) # 글자 색은 빨갛게, 이탤릭, 두껍게 적용
b1.font = Font(color="CC33FF", name="Arial", strike=True) # 폰트 : Arail, 취소선 적용
c1.font = Font(color="0000FF", size=20, underline="single")    # 글자 크기를 20으로 밑줄 적용

# 테두리 적용
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

a1.border = thin_border
b1.border = thin_border
c1.border = thin_border

# 90 점 넘는 셀에 대해서 초록색으로 적용
for row in ws.rows:
    for cell in row:
        if cell.column == 1:
            continue

        if isinstance(cell.value, int) and cell.value > 90:
            cell.fill = PatternFill(fgColor="00FF00", fill_type="solid")
            cell.font = Font(color="FF00000") # 폰트 색상 변경


wb.save("sample_stype.xlsx")
